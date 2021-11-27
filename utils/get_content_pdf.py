import requests
from bs4 import BeautifulSoup
import numpy as np
from openpyxl import workbook
from openpyxl import load_workbook
import traceback
import os
import time
from selenium import webdriver
import csv
import PyPDF2
import tabula
import concurrent.futures
import re





#The function to get the data from the PDF data of the SPK is not yet implemented.
# The name of the accounts and the TAG is obtained directly from the name of the PDF.
# What needs to be implemented is an alg. that reads the tables and filters out price points with the invoices in the dat.
#The terms are, as I have seen so far, not identical to those of the VRB, but it can be well oriented.

###Sparkassen###

def get_name_sp(file):
    #os.chdir("filesSp")
    pdfFileObj = open(file, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    for page in range(pdfReader.numPages):
        pageObj = pdfReader.getPage(page)
        text = pageObj.extractText()
        start = text.find("Kontobezeichnung")
        end =  text.find("Datum")
        #print("Text:",text,start,end,"\n")
        if start == -1 or end == -1:
            continue
        else:
            name = text[start+17:end]
            # bei text[start+18]kann manchmal ein : sein, also gegenebfalls rausnehmen mit extra if
            return [name,page+1]
    return 0
#print(getNameSp("kasseler-sparkassegirokomfort.pdf"))

def get_content_pdf_sp(file):
    pdfFileObj = open(file, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    for page in range(pdfReader.numPages):
        pageObj = pdfReader.getPage(page)
        text = pageObj.extractText()
        start = text.find("Kontoführungmonatlich")
        end =  text.find("Jährliche Gesamtentgelte")
        if start == -1 or end == -1:
            continue
        else:
            content = text[start+21:end].replace(",",".")
            retcontent = ""
            for digit in range(len(content)):
                if content[digit].isdigit() or content[digit] == ".":
                    retcontent = retcontent + content[digit]
            return float(retcontent)
    return float(0)

def getcontentCsvSp(file):

    return 0



def mainSp():

    return 0


###Volksbanken###

#Funktion schreibt alle Namen von Konten aus einer PDF Datei heraus und gibt zusätzlich die Seite auf der der Name gefunden wurde an
def get_name(file):
    #os.chdir("filesVr")
    end = False
    names = []
    cnt = 1
    while not end:
        res = findNameInPdf(file,cnt)
        if res != None:
            names.append(res)
        else:
            end = True
        cnt += 1
    return names


#Take file and return iter name of account.
def get_name_in_pdf(file,iter):
    #Öffne PDF
    pdfFileObj = open(file, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

    cntFound = 0
    #iteriere über alle Seiten der PDF
    for page in range(pdfReader.numPages):
        # if page < iter-1:
        #     continue
        #Seite mit der nummer "page" laden
        pageObj = pdfReader.getPage(page)
        try:
            #Text der Seite nehmen
            text = pageObj.extractText()
        except:
            #Bei Fehler abbrechen
            return None
        #Die PDFs sind so formatiert, dass du zwischen den Begriffen "Kontobezeichnung" und "Datum" den Namen hast.
        start = text.find("Kontobezeichnung")
        end =  text.find("Datum")
        #.find() gibt -1 zurück wennn nichts in dem angegebenen Objekt gefunden wurde, in diesem Fall wird mit "continue" die Schleife von Vorner, mit der nächsten Seite, fortgesetzt
        if start == -1 or end == -1:
            continue
        else:
            #"Konotobezeichnung" ist 17 Zeichen lang, manchmal ist ein ":"
            name = text[start+17:end].strip(":")
            #Teilweise laden die Banken "___________" als Kontobezeicnung hoch, das umgehe ich hier
            if "__" in name:
                continue

            cntFound += 1
            #Ich das mache ich um sagen zu können den wievielten Namen ich aus der Datei haben möchte, geht bestimmt auch einfacher
            if cntFound == iter:
                return [name,page+1]
    return None



#Findet den Start einer Preisangabe in der PDF Datei
def find_start(content):
    #Mit KOntoführung beginnt die Preisanagebe immer
    if "Kontoführung" in content:
        #dann muss ich jedoch zwischen verschiedenen Fällen unterscheiden
        if "onatlich" in content:
            return [content.find("onatlich")+8,1]
        elif "rund" in content:
            return [content.find("rund")+4,1]
        elif "rundpreis" in content:
            return [content.find("rundpreis")+9,1]
        elif "auschale" in content:
            return [content.find("auschale")+8,1]
            #Wenn der Preis quartalsweise angegeben ist muss ich den Preis auf monatasbasis normieren
        elif "uartal" in content:
            return [content.find("uartal")+6,1/3]
        elif "_" in content:
            return [content.find("_"),1]
        #Manchmal wird "kein turnusmäßiges Entgelt" benutzt, dann gebe ich 0 zurück
        elif "turnusmäßiges" in content:
            return [0,1]

    else:
        return [-1,1]


#Die Preisangabe endet mit der Angabe einer Einheit. "\u20ac" ist das Euro Symbol.
#Ein Problem was ich hier noch überprüfen muss ist, ob diese Funktion nicht zu weit geht. Also ob ich wirklich die erste Einheitsangabe treffe und so den richtigen Preis herausfiltere.
def find_end(content,start):
    if "EUR" in content:
        return content[start:].find("EUR")
    elif "Euro" in content:
        return content[start:].find("Euro")
    elif "\u20ac" in content:
        return content[start:].find("\u20ac")
    else:
        return -1


#Zahl aus String schreiben die in einem bestimmten Format ist, z.b. X.X wobei X beliebige Dezimalzahlen sind
def find_num(str):
    cont = str.replace(",",".")
    try:
        return re.findall("\d+\.\d+",cont)[0]
    except:

        return 0
    #traceback.print_exc()
    # retcontent = ""
    # for digit in range(len(content)):
    #     if content[digit].isdigit() or content[digit] == ".":
    #         retcontent = retcontent + content[digit]
    # return float(retcontent)*start[1]
    #
    # if retcontent[-5:-4].isdigit():
    #     return float(retcontent[-5:])*start[1]
    # return float(retcontent[-4:])*start[1]


#Funktion die Content einer PDF Datei nimmt, implementiert durch pageObj.extractText().
def get_content_pdf_vr(file,currentPage,nextPage):
    #os.chdir("filesVr")
    pdfFileObj = open(file, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    for page in range(currentPage-1,nextPage-2):

        pageObj = pdfReader.getPage(page)
        text = pageObj.extractText()

        start = findStart(text)
        #print("start",start)
        if start == None:
            #print("yes")
            continue
        elif start[0] == -1:
            continue
        elif start[0] == 0:
            break

        end = findEnd(text,start[0])

        if end == -1:
            return float(0)

        else:
            content = text[start[0]:start[0]+end]
            return findNum(content)


    return float(998)



#Teilweise sind die PDF Dateien falsch formatiert von den Banken hochgeladen worden. Dann nehme ich die rechenaufwändigere CSV methode um die Preise zu bekommen
def get_content_csv_vr(file,currentPage,nextPage):
    #os.chdir("filesVr")
    # Wenn ich nur noch eine Seite übrig habe
    if currentPage == nextPage:
        tabula.convert_into(file, "output.csv", output_format="csv", pages="{0}-{1}".format(currentPage,nextPage))
    else:
        #Die Seite die zu einem Konto gehört in CSV umwandeln
        tabula.convert_into(file, "output.csv", output_format="csv", pages="{0}-{1}".format(currentPage,nextPage-1))

    with open("output.csv", 'r') as csvfile:
        reader = csv.reader(csvfile, delimiter=' ', quotechar='|')

        for row in reader:
            content = ''.join(row)

            start = findStart(content)
            #print(content)
            #Wenn Kontoführung aber kein match mit weiteren Begriffen
            if start == None:
                continue

            #Wenn Kontoführung nicht zu finden ist
            elif start[0] == -1:
                continue

            #print(content)
            end = findEnd(content,start[0])
            #print("End",end)
            #Wenn kein Symbol für das Ende zu finden ist, also wenn zwar monatlich steht aber kein Euro symbol. Dann habe ich angenommen, dass der Preis 0 ist. Ich habe bis jetzt keinen Fall gefunden wo das nicht funktioniert.
            if end == -1:
                return float(0)

            else:
                content = content[start[0]:start[0]+end]
                return findNum(content)

    #Im Tool wird nach 404 gefiltert, das entspricht den Konten für die ich keine Preise durch das Tool finden kann. Gerade sind das 2.6% der Konten.
    return float(404)

#print(getcontentCsvVr("echterdinger-bank.pdf",5,9))






#Alte FUnktion
def main_vr():
    wb = load_workbook("EntgeltinformationVr.xlsx")
    sheet = wb[wb.sheetnames[0]]
    os.chdir("filesVr")
    filesPdf = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]

    cnt = 0
    cntcsv = 0

    for file in filesPdf:
        print(file)
        names = getName(file)
        for name in range(len(names)):
            try:
                content = getcontentCsvVr( file,  names[name][1], names[name+1][1] )
                if content == 998:
                    content = getcontentPdfVr(file,  names[name][1], names[name+1][1])
                sheet.cell(row = cnt+2, column = 1).value = file.replace(".pdf","")
                sheet.cell(row = cnt+2, column = 2).value = names[name][0]
                sheet.cell(row = cnt+2, column = 3).value = content
                cnt += 1
                print(names[name][0]," ",content)
            except:
                if name + 1 == len(names):
                    sheet.cell(row = cnt+2, column = 1).value = file.replace(".pdf","")
                    sheet.cell(row = cnt+2, column = 2).value = names[name][0]
                    sheet.cell(row = cnt+2, column = 3).value = getcontentCsvVr(file, names[name][1], PyPDF2.PdfFileReader(open(file, 'rb')).numPages)
                    cnt += 1
                else:
                    print("ERR begin")
                    traceback.print_exc()
                    print("ERR end")

                    # if not there 0
        cntcsv += 1
        print(len(filesPdf)-cntcsv)
    wb.save("d:/7357/Desktop/tool/test.xlsx")
    wb.close()
    return 0


def extract_data(file):
    #os.chdir("filesVr")
    #print("Curretnly using this file:",file)
    # if isEmpty(file.replace(".pdf",".csv")) :
    #     print("empty")
    #     return [[file.replace(".pdf",""),"",float(888)]]
    names = getName(file)
    #print("names:",names)
    #ret = [[0,0,0]]*len(names)
    ret = []
    cnt = 0
    for name in names:

        if cnt + 1 == len(names):
            #Letzter Name
            content = getcontentPdfVr(file, name[1], PyPDF2.PdfFileReader(open(file, 'rb')).numPages)
            #Speicher ich ab um mir eine IF Bed. für den Fall dass über extract.text nichts gefunden wird
            end = PyPDF2.PdfFileReader(open(file, 'rb')).numPages

        else:
            content = getcontentPdfVr( file, name[1], names[cnt+1][1] )
            end = names[cnt+1][1]
            #print("from PDF")
        #Fall dass über die PDF nichts gefunden wird
        if content == float(998):
            #Hier setze ich dann das Ende ein.
            content = getcontentCsvVr(file, name[1], end)
            #print("from csv")

        # ret[cnt][0] = file.replace(".pdf","")
        # ret[cnt][1] = name[0]
        # ret[cnt][2] = content
        # Ich habe den TAG und den Namen des Kontos mit "qq" in der Dateibezeichnung getrennt.
        pos = file.find("qq")
        instName = file.replace(".pdf","")
        ret.append([instName[:pos],name[0],content])
        #print("This content was written:",ret)
        cnt += 1
        #print("content:",ret,"\n")

    return ret

#print(extractData("ersbruck.pdf"))
#Die Funktion ist logisch in 2 Teile aufzusplitten. Einmal schreibe ich aus der Excel alle pdf Tags raus.
#Dann gehe ich die Liste der Entgelt PDFS durch und streiche die welche in der Tag List sind raus
#So bleiben dann nur die PDFS übrig welche noch nicht in der Excel sind.
def remove_element(listPdf,sheet):
    isIn = []
    end = False
    cnt = 2
    while not end:
        # Gibt Inhalt einer Zelle A2, A3, A4,.. an
        data = sheet["A{}".format(cnt)].value
        #print(data,cnt)
        #Wenn data = None dann ist zu Ende
        if data != None:
            cnt += 1
            if data in isIn:
                continue
            else:
                isIn.append(data)

        else:
            end = True
    #print(isIn)
    for remPdf in isIn:
        try:
            listPdf.remove(remPdf+".pdf")
            #print("yes",remPdf)
        except:
            0
            #print("no",remPdf)
    return [listPdf,cnt]



#Eigentliche Funktion. Ich threade das herauszuschreiben, deshalb ist die CPU vom PC sehr ausgelastet wenn du das Programm laufen lässt.
#Man kann hier statt threading sicherlich auch multi-proccessing (https://www.machinelearningplus.com/python/parallel-processing-python/) viel an Rechenleistung rausholen.
def main_vr_Thread():
    wb = load_workbook("EntgeltinformationVr.xlsx")
    sheet = wb[wb.sheetnames[0]]
    os.chdir("filesVr")
    filesPdf = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]
    rem = removeElement(filesPdf,sheet)
    filesPdf = rem[0]
    #return filesPdf
    #Row in der ich in der Excel starte zu schreibe, sodass ich keine Daten überschreibe
    rowS = rem[1]


    cntiter = 0

    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Start the operations and mark each future with its URL
        data_to_content = {executor.submit(extractData, file): file for file in filesPdf}

        for future in concurrent.futures.as_completed(data_to_content):
            try:
                result = future.result()
                #print(result)
            except:
                print("Fehler in Auswertung")
                traceback.print_exc()
                continue
            #print("result: ",result)

            for val in result:
                sheet.cell(row = rowS, column = 1).value = val[0]
                sheet.cell(row = rowS, column = 2).value = val[1]
                sheet.cell(row = rowS, column = 3).value = float(val[2])
                rowS += 1

            cntiter += 1
            #if cntiter%20==0:
            print(len(rem[0])-cntiter)

    wb.save("d:/7357/Desktop/tool/EntgeltinformationVr.xlsx")
    wb.close()
    return 0




#Alte funktionen die ich nur einmal gebraucht habe um die Namen der Konten herauszuschreiben.
def extractDataNames(file):
    if isEmpty(file.replace(".pdf",".csv")):
        return [[file.replace(".pdf",""),"",float(999)]]
    names = getName(file)
    ret = []
    for name in names:
        ret.append([file.replace(".pdf",""),name[0]])
    return ret

def mainvrThreadNames():
    wb = load_workbook("NameVR.xlsx")
    sheet = wb[wb.sheetnames[0]]
    os.chdir("filesVr")
    filesPdf = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]

    rowS=2
    cntiter = 0

    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Start the operations and mark each future with its URL
        data_to_content = {executor.submit(extractDataN, file): file for file in filesPdf}

        for future in concurrent.futures.as_completed(data_to_content):
            try:
                result = future.result()
            except:
                traceback.print_exc()
                continue
            #print("result: ",result)

            for val in result:
                sheet.cell(row = rowS, column = 1).value = val[0]
                sheet.cell(row = rowS, column = 2).value = val[1]
                rowS += 1

            cntiter += 1

            print(len(filesPdf )-cntiter)

    wb.save("d:/7357/Desktop/tool/NameVR.xlsx")
    wb.close()
    return 0


start = time.time()
print(mainvrThread())
end = time.time()
print("Runtime was",round(end-start,1))
