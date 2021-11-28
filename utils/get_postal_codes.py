import requests
from requests import get
from bs4 import BeautifulSoup

import urllib.request
import numpy as np
import time
import pandas as pd
import json
import os
import traceback
from openpyxl import workbook
from openpyxl import load_workbook


# Takes URL and returns list with specific data points of the page
def get_data_sp(url, proxy=None):
    dataArraySp = np.array(["", "", 0, 0, 0, 0], dtype="<U132")
    try:
        response = requests.get(url, proxy)
    except:
        response = False
    if str(response) != "<Response [200]>":
        print("{0}This url didnt time work: {1}".format(str(response), url))
        return np.array(["", "", 0, 0, 0, 0])
    soup = BeautifulSoup(response.text, "html.parser")
    try:
        # Suche nach dem Skript Tag, gehe an bestimmte Stelle vom Array, lade JSON Objekt
        data = json.loads(str(soup.findAll('script')[7])[35:-9])
    except:
        # Manchmal sind die Daten an einer anderen Stelle gespeichert..
        data = json.loads(str(soup.findAll('script')[9])[35:-9])
    # Name Institut
    dataArraySp[0] = data[0]["name"]
    # Name Filiale, manchmal nicht vorhanden
    try:
        dataArraySp[1] = data[1]["name"]
    except:
        # Teilweise steht der Name der Filiale nicht in dem Textblock aus dem ich die restlichen Informationen ziehe.
        dataArraySp[1] = data[0]["name"]
    # PLZ
    dataArraySp[2] = data[0]["address"]["postalCode"]
    try:
        # longitude
        dataArraySp[3] = data[0]["geo"]["longitude "]
        # latitude
        dataArraySp[4] = data[0]["geo"]["latitude"]
    except:
        # Falls nicht dort Mitte von DE, kam aber noch nicht vor
        dataArraySp[3] = 10.447683
        # latitude
        dataArraySp[4] = 51.163361
    try:
        # Die Blz ist in einem Button zur Sparkasse versteckt, deshalb suche ich nach der Klasse mit allen "data-blz" Attributen
        dataArraySp[5] = soup.find_all("button", attrs={"data-blz": True})[0]["data-blz"]
    except:
        dataArraySp[5] = 0

    return dataArraySp


def get_data_vr(url, proxy=None):
    dataArrayVr = np.array([0, "", 0], dtype="<U140")
    try:
        response = requests.get(url, proxy)
    except:
        response = False
    if str(response) != "<Response [200]>":
        print("{1} this url didnt work: {2}".format(str(response), url))
        return np.array([0, "", 0])
    soup = BeautifulSoup(response.text, "html.parser")
    data = soup.findAll('div', 'text')
    # Institut
    dataArrayVr[0] = str(data[0])[str(data[0]).find('BLZ') + 5:str(data[0]).find('BLZ') + 13]
    # Filiale
    dataArrayVr[1] = data[0].find('h1', 'branch-office-header__headline').get_text()
    # PLZ
    dataArrayVr[2] = data[0].findAll('span')[1].get_text()
    # Koordinaten sind leider nicht in den HTML Inf. enthalten, werden aber in der Excel nachgeschlagen.
    return dataArrayVr


# Hauptprogramm SPK. Kann sein, dass du geblockt wirst, durch das Threading werden viele Anfragen gestellt. In dem
# Fall nimmst du entweder Proxy oder wechselst zum 2.0 Progromm.



# Non Thread Version
def main_sp_2(proxy=None):
    file = open("UrlSpFiliale.txt", "r")
    cnt = 6
    wb = load_workbook("Postalcode Tool.xlsm", read_only=False, keep_vba=True)
    sheet = wb[wb.sheetnames[1]]
    lenUrl = len(open("UrlSpFiliale.txt").readlines())
    for iter in range(lenUrl):

        data = get_data_sp(file.readline().replace('\n', ''))

        try:
            sheet.cell(row=cnt, column=1).value = data[0]
            sheet.cell(row=cnt, column=2).value = data[1]
            sheet.cell(row=cnt, column=3).value = float(data[2])
            sheet.cell(row=cnt, column=5).value = float(data[3])
            sheet.cell(row=cnt, column=4).value = float(data[4])
            # BLZ
            sheet.cell(row=cnt, column=7).value = float(data[5])
            cnt += 1

        except:
            traceback.print_exc()
            continue
        print(lenUrl - cnt)
        if "00" in str(cnt):
            print("save")
            # Zwischenstand speichern
            wb.save("Postalcode Tool.xlsm")

    wb.save("Postalcode Tool.xlsm")
    file.close()
    return 0


