import requests
from bs4 import BeautifulSoup
import numpy as np
from openpyxl import workbook
from openpyxl import load_workbook
import traceback
import os
import time
import tabula
import csv
import PyPDF2
import re


#Loading of the price notices of SP.
def get_preisaushang_pdf_sp():
    #URL der Filiale mit Pfad zu den Infos öffnen
    file = open("url_sp_entgelt.txt","r")
    #Für den Tag des Entgeltes öffnen
    file2 = open("url_sp_Institut.txt","r")
    cnt = 0
    for num in range(len( open(   "url_sp_entgelt.txt").readlines(    )  ) ):
        url = file.readline().replace('\n','')
        address = file2.readline().replace('\n','')
        name = address[11:-4].replace(".","")
        #print("This url is tested: {}".format(url))


        response = requests.get(url)
        print(response)
        soup = BeautifulSoup(response.text, "html.parser")
        href = soup.findAll('a')

        try:
            for i in href:
                if "preisaushang.pdf" in str(i.get('href')):
                    data = requests.get(   address.replace("'","") + str(i.get('href'))    ,    allow_redirects=True)
                    open('D:/7357/Desktop/tool/filesSp/{}.pdf'.format(name), 'wb').write(data.content)

                #print("This did not work: {}".format(name))
        except:
            print("no href" + " "+str(cnt)+" "+ address + str(i.get('href')) )
        cnt += 1
        print(len(open("UrlSpEntgelt.txt").readlines())-cnt,"left")
    file.close()
    file2.close()
    return 0


#Laden der Entgeltinformationen von Sparkassen
def get_entgelt_sp():
    file = open("url_sp_entgelt.txt","r")
    file2 = open("url_sp_Institut.txt","r")

    cnt = 0
    for num in range(len( open(   "url_sp_entgelt.txt").readlines(    )  ) ):
        url = file.readline().replace('\n','')
        address = file2.readline().replace('\n','')


        #Der Name wird als Tag in der Bezeichnung der PDF-Datei benutzt. Das mache ich so um später im Tool die Zuordnung zu den PLZ Bereichen haben
        if address[:5] == "https":
            name = address[12:-4].replace(".","")
        else:
            name = address[11:-4].replace(".","")


        try:
            response = requests.get(url)
        except:
            print("failed to connect",url)
            continue

        soup = BeautifulSoup(response.text, "html.parser")
        # Suche nach Begriffen die "ntgeltinfo" im href haben. Die Schreibweise wurde gewählt um Entgeltinfo/entgeltinfo/Entgeltinformation/entgeltinformaiton zu treffen.
        href = soup.findAll('a',href = re.compile("ntgeltinfo"))#,'target':'_blank'})
        #print(len(href))
        #In jeder Iteration wird found aus falsch gesetzt
        found = False

        for i in href:

            #print("yes")
            #Zwischen den beiden Punkten sollte die Bezeichnung des Kontos liegen.
            start = i.get('href').find('ntgeltinfo')
            end = i.get('href').find('.pdf?')
            #print("here",name)
            try:
                #Daten der PDF Datei abrufen
                data = requests.get(   address.replace("'","") + str(i.get('href'))    ,    allow_redirects=True)
                #PDF content in das Verzeichnis schreiben.
                # Ich habe mich für "qq" als Trennung zwischen Tag und Name des Kontos entschieden, da die Kombination von Buchstaben an sich nicht natürlich auftreten sollte.
                open('D:/7357/Desktop/tool/filesSp/{}.pdf'.format(name+"qq" +str(i.get('href'))[start+18:end].replace("/","") ), 'wb').write(data.content)
                found = True
                cnt+=1


            except:
                traceback.print_exc()

        #print(cnt)
        #Falls ich mit "ntgeltinfo" nichts gefunden habe suche ich weiter nach abweichenden Bezeichnungen
        if not found:
            #Hiermit bekomme ich auf jeden Fall die Entgeltinfos, leider auch alle anderen PDFs auf der Internetseite. Da diese dann aber später nicht vom getContent Programm ausgewertet werden können ist das nicht so schlimm.
            href = soup.findAll('a',href = re.compile("preise-leistungen"))
            for i in href:
                start = i.get('href').find('preise-leistungen')
                end = i.get('href').find('.pdf?')
                try:
                    data = requests.get(   address.replace("'","") + str(i.get('href'))    ,    allow_redirects=True)
                    open('D:/7357/Desktop/tool/filesSp/{}.pdf'.format(name+"qq" +str(i.get('href'))[start+37:end].replace("/","") ), 'wb').write(data.content)
                    cnt+=1


                except:
                    traceback.print_exc()



    print("With ",len(open("url_sp_entgelt.txt").readlines()),"Websites",cnt, "PDF were collected")
    file.close()
    file2.close()

    return 1





#Alte Funktion um Überblick der Kontobezeichnungne SPK zu bekommen
def get_name():
    wb = load_workbook("test.xlsx")
    sheet = wb[wb.sheetnames[0]]
    cnt = 1
    for num in range(len( open(   "UrlSpEntgelt.txt").readlines(    )  ) ):
        url = file.readline().replace('\n','')
        address = file2.readline().replace('\n','')


        if address[:5] == "https":
            name = address[12:-4].replace(".","")
        else:
            name = address[11:-4].replace(".","")
        #print("This url is tested: {}".format(url))


        response = requests.get(url)

        soup = BeautifulSoup(response.text, "html.parser")
        href = soup.findAll('a')

        try:
            for i in href:
                if "entgeltinformation" in str(i.get('href')) and ".pdf" in str(i.get('href')) :
                    start = i.get('href').find('entgeltinformation')
                    end = i.get('href').find('.pdf?n=true')
                    leng = len( i.get('href') )

                    sheet.cell(row = cnt, column = 1).value = name
                    sheet.cell(row = cnt, column = 2).value = str(i.get('href'))[start+19:end]
                    cnt +=1

        except:
            print("Error in " + " "+ address + str(i.get('href')) )

        print(len(open("UrlSpEntgelt.txt").readlines())-cnt,"left")



    wb.save("test.xlsx")
    return 0





###Volksbanken###


def get_preisaushang_pdf_vr():
    #Wechsel in das richtige Verzeichnis
    os.chdir("filesVr")
    #Alle PDF Dateien in eine Liste schreiben
    files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]
    #Wechsel zurück damit ich im nächsten Schritt die txt Datei öffnen kann
    os.chdir("..")

    file = open("UrlVrInstitut.txt","r")
    iter = 0
    for num in range(len( open("UrlVrInstitut.txt","r").readlines(    )  ) ):

        url = file.readline().replace("\n","")
        if num < iter :
            continue
        end = url.find(".de")
        if url[:5] == "https":
            name = url[12:end]
        else:
            name = url[11:end]
        #Wenn ich von der Bank mit dem Tag schon die PDF habe dann überspringe diese URL
        if "{}.pdf".format(name) in files:
            continue


        found = False
        cnt = 0
        #Super uneffizient, hier findet man zu 100% eine bessere Möglichkeit. Ich hatte nur keine Zeit mehr. Die URL Anhänge sind wenigstens schon nach wichtigkeit sortiert.
        tryAr = ['/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html','/Service/rechtliche-hinweise/pflichtinformationen.html','/service/rechtliche-hinweise/Pflichtinformationen/entgeltinformationen.html','/service/pflichtinformationen/entgeltinformationen.html','/service/Pflichtinformationen/entgeltinformationen.html','/service/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html','/service/zahlungskontengesetz.html/service/entgeltinformationen.html ','/service/pflichtinformationen/entgeltinformationen.html#parsys_seitenkopf','/pflichtinformationen/entgeltinformationen/zahlungskontengesetz-zkg/','/service1/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html','/service/rechtliche-hinweise/pflichtinformationen1/entgeltinformationen.html','/service0/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html','/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html','/ihre-bank/rechtliches/pflichtinformationen/c1090.html#9836','/service/rechtliche_hinweise/pflichtinformationen/entgeltinformationen.html','/service/rechtliche-hinweise/pflichtinformationen0/entgeltinformationen.html','/service/Pflichtinformationen.html/pflichtinformationen/','/service/sitemap/Zahlungskontengesetz.html#parsys_seitenkopf','/service1/pflichtinformationen/entgeltinformationen.html','/ueber_uns/bekanntmachungen/Entgelttransparenz.html','/service/pflichtinformationen.html','/Service/pflichtinformationen/entgeltinformationen.html','/service/RechtlicheHinweise/pflichtinformationen/entgeltinformationen.html','/service/rechtlicheHinweise/pflichtinformationen/entgeltinformationen.html','/service1/rechtliche-hinweise/pflichtinformationen0/entgeltinformationen.html','/service/informationen-zur-entgelttransparenz.html','/service/rechtliche-inhalte/pflichtinformationen/zahlungskontengesetz.html','/service/zahlungskontengesetz--zkg-.html','/index.php/zkg.html','/service/nutzungsbedignungen1/entgeltinformationen.html','/module.php5?mod=vorlagen&fid=7&ident=32','/service/rechtliche-hinweise/pflichtinformationen/Zahlungkontengesetz1.html','/service/informationspflicht/zahlungskontengesetz.html','/rechtliches/pflichtinformationen/entgeltinformationen.html']
        

        try:
            while not found:
                #print("this url",url)
                tryUrl = url+tryAr[cnt]
                response = requests.get(tryUrl)
                #print("trying",url+tryAr[cnt])
                if str(response) != "<Response [200]>":
                    cnt += 1
                else:
                    found = True


        except:
            traceback.print_exc()
            print("No success with current tryAr",name)
            print(num)
            continue

        #suche nach den selben tags wie oben, wenn dann mehr als einer drin ist engage in das neue, sonst standard abgreifen
        soup = BeautifulSoup(response.text, "html.parser")
        href = soup.findAll('a',href = re.compile("ntgeltinfo"))#,'target':'_blank'})


        #In jeder Iteration wird found aus falsch gesetzt
        foundEnt = False

        for i in href:

            #Zwischen den beiden Punkten sollte die Bezeichnung des Kontos liegen.
            start = i.get('href').find('ntgeltinfo')
            end = i.get('href').find('.pdf?')
            #Bei VR Banken sind die Entgeltinfos entweder zusammengefasst in einer Datei oder einzeln auf den Websites. Im gesammelten Fall ist len(href)==1 und ich gebe den passenden Namen. Sonst ist der Algorithmus der selbe wie auch bei SPK.
            if len(href)==1:
                accountName = "kombiniert"
            else:
                accountName = str(i.get('href'))[start+18:end].replace("/","")
            try:

                #Die VR Banken habe für die Angabe des HREFs 2 verschiedene Logiken, manchmal ist die URL des Instituts schon mit drin, manchmal nicht
                address = str(i.get('href')  )
                if address[:4] != "http":
                    address = url+address

                #Daten der PDF Datei abrufen
                data = requests.get(   address    ,    allow_redirects=True)
                #PDF content in das Verzeichnis schreiben.
                # Ich habe mich für "qq" als Trennung zwischen Tag und Name des Kontos entschieden, da die Kombination von Buchstaben an sich nicht natürlich auftreten sollte.
                open('D:/7357/Desktop/tool/filesVr/{}.pdf'.format(name+"qq" + accountName), 'wb').write(data.content)
                foundEnt = True
                cnt+=1
            except:
                traceback.print_exc()

            if not foundEnt:
                print("not found",name)


    file.close()
    return 0
print(getPreisaushangPDFVR())


#Alte Funktion
def getPreisaushangUeberUrl():
    os.chdir("filesVr")
    files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]
    os.chdir("..")
    present = len(files)
    required = len( open(   "UrlVrInstitut.txt").readlines(    )  )
    cnt = 0
    file = open("UrlVrInstitut.txt","r")
    foundcnt = 0
    for num in range(required):

        url = file.readline().strip('\n')
        end = url.find(".de")
        if url[:5] == "https":
            name = url[12:end]
        else:
            name = url[11:end]
        if "{}.pdf".format(name) in files:
            #print("is in:",num)
            cnt+=1
            foundcnt += 1
            continue
        print("is searched:",num)


        found = False

        response = requests.get(url+"/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html") #   "/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html"    "/service/entgeltinformationen.html" "/service/pflichtinformationen/entgeltinformationen.html#parsys_seitenkopf" "/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html"
        if str(response) != "<Response [200]>":
            print("Error in starting ",url)
            cnt+=1

            continue
        soup = BeautifulSoup(response.text, "html.parser")
        href = soup.findAll('a')
        if href == []:
            print("empty",cnt)
        for i in href:

            try:
                if "Entgeltinformationen" in str(i.get_text()):# and "pdf" in str(i.get('href')):
                    found = True
                    print(cnt,found,i.get('href'))
                    pdfUrl = str(i.get('href'))
                    if pdfUrl[:4] != "http":
                        pdfUrl = url[:end+3]+pdfUrl
                    #print(pdfUrl)
                    foundcnt += 1
                    data = requests.get(pdfUrl ,  allow_redirects=True)
                    open('D:/7357/Desktop/tool/filesVr/{}.pdf'.format(name), 'wb').write(data.content)
                    cnt+=1

                    break

                else:
                    0
            except:
                0
        if found != True:
            print( "didnt find {}".format(url))

    file.close()
    print("Left:",required-foundcnt)
    return 0


#Alte Funktion, kann dazu genutzt werden Eingaben in Suchfeldern zu machen
def getPreisaushangWebdriver():
    os.chdir("filesVr")
    files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.pdf')]
    os.chdir("..")
    present = len(files)
    required = len( open(   "VrUrlInst.txt").readlines(    )  )
    cnt = 0
    file = open("VrUrlInst.txt","r")
    driver = webdriver.Chrome()
    for num in range(required):

        url = file.readline().strip('\n')
        end = url.find(".de")
        if url[:5] == "https":
            name = url[12:end]
        else:
            name = url[11:end]
        if "{}.pdf".format(name) in files:
            #print("is in:",num)
            cnt+=1
            continue
        print("is searched:",num)


        found = False

        response = requests.get(url+"/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html") #"/service/entgeltinformationen.html" "/service/pflichtinformationen/entgeltinformationen.html#parsys_seitenkopf" "/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html"


        if str(response) != "<Response [200]>":
            print("Error in starting ",url)
            cnt+=1
            continue
        driver.get(url+"/service/rechtliche-hinweise/pflichtinformationen/entgeltinformationen.html")
        input()
    #     time.sleep(5)
    #     soup = BeautifulSoup(driver.page_source, "html.parser")
    #     href = soup.findAll('a')
    #     for i in href:
    #
    #         try:
    #             if "Entgeltinformationen" in str(i.get_text()):# and "pdf" in str(i.get('href')):
    #                 found = True
    #                 print(cnt,found,i.get('href'))
    #                 pdfUrl = str(i.get('href'))
    #                 if pdfUrl[:4] != "http":
    #                     pdfUrl = url[:end+3]+pdfUrl
    #                 #print(pdfUrl)
    #                 data = requests.get(pdfUrl ,  allow_redirects=True)
    #                 open('D:/7357/Desktop/tool/filesVr/{}.pdf'.format(name), 'wb').write(data.content)
    #                 cnt+=1
    #                 break
    #
    #             else:
    #                 0
    #
    #         except:
    #             0
    #         time.sleep(5)
    #     if found != True:
    #         print( "didnt find {}".format(url))
    # driver.quit()

    file.close()
    print("Left:",required-present-cnt)
    return 0
