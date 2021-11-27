import requests
from bs4 import BeautifulSoup
import numpy as np
from openpyxl import workbook
from openpyxl import load_workbook
import traceback
import os
import time
import tabula
import csv
import PyPDF2
import re


# Function that writes the commercial register information in CSV file
def main_handels_vr(url_txt):
    # "UrlVrInstitute.txt"
    file = open(url_txt, "r")
    file2 = open("commercial_register_vr.csv", "w+")
    cnt = 0
    for num in range(len(open(url_txt).readlines())):
        url = file.readline().replace('\n', '')
        if url[:5] == "https":
            name = url[12:-3].replace(".", "")
        else:
            name = url[11:-3].replace(".", "")
        file2.write("{}".format(name))
        # print("This url is being tested: {}".format(url))
        response = requests.get(url)
        if str(response) != "<Response [200]>":
            print("Error, this didn't work", url)
            continue
        soup = BeautifulSoup(response.text, "html.parser")
        content = soup.findAll('a')

        try:
            for i in content:
                text = i.get_text().replace("\n", "")
                if "Impressum" in str(i.get_text()):
                    response = requests.get(i.get("href"))
                    soup = BeautifulSoup(response.text, "html.parser")
                    handels = soup.findAll('p')
                    foundReg = False
                    foundGer = False
                    end = False
                    for y in handels:
                        if foundReg:
                            text = y.get_text()
                            file2.write(";{}".format(re.findall("\d+", text)[0]))
                            foundReg = False
                            continue
                        if foundGer:
                            text = y.get_text()
                            file2.write(";{}\n".format(text.replace("Amtsgericht", "")))
                            end = True
                            foundGer = False
                            break
                        if "Genossenschaftsregister" in str(y.get_text()):
                            foundReg = True
                            if len(str(y.get_text())[23:]) > 2:
                                file2.write(";" + re.findall("\d+", str(y.get_text()))[0])
                                foundReg = False

                            continue
                        if "Registergericht" in str(y.get_text()):
                            foundGer = True
                            if len(str(y.get_text())[15:]) > 2:
                                file2.write(";" + str(y.get_text())[15:].replace(":", "").replace("\n", "").replace(
                                    "Amtsgericht", ""))
                                foundGer = False
                            continue
                    if not end:
                        file2.write("\n")
                    break
                # print("This did not work: {}".format(name))
        except:
            print(url, "didnt work")

    file.close()
    file2.close()
    return 0


def main_handels_sp(url_txt):
    # Für die Url der Entgeltinfos öffnen
    file = open(url_txt, "r")
    # Für den Tag des Entgeltes öffnen
    file2 = open("commercial_register_sp.csv", "w+")
    cnt = 0
    for num in range(len(open(url_txt).readlines())):
        url = file.readline().replace('\n', '')
        if url[:5] == "https":
            name = url[12:-3].replace(".", "")
        else:
            name = url[11:-3].replace(".", "")
        file2.write("{}".format(name))
        # print("This url is tested: {}".format(url))
        response = requests.get(url + "de/home/toolbar/impressum.html?n=true&stref=footer")
        if str(response) != "<Response [200]>":
            print("didnt work", url)
            continue
        soup = BeautifulSoup(response.text, "html.parser")
        content = soup.findAll("td", "left")

        try:
            found = False
            for i in content:
                text = i.get_text().replace("\n", "")
                if "sgericht" in str(text) or "HR" in str(text):
                    # print(text)
                    cnt += 1
                    found = True
                    break
            if not found:
                print("didnt find", url)

        except:
            print(url, "didnt work")
        # if cnt == 10:
        #     break

    file.close()
    file2.close()
    return 0


def find_tag_vr(url_txt):
    file = open(url_txt, "r")
    file2 = open("url_vr_tag.csv", "w+")
    for url in range(len(open(url_txt).readlines())):
        try:
            response = requests.get(file.readline().replace('\n', ''))
        except:
            print("This didn't work: ", url)
            continue

        soup = BeautifulSoup(response.text, "html.parser")
        href = soup.findAll("span", "module-linklist__title")
        data = soup.findAll('div', 'text')
        try:
            branch = str(data[0])[str(data[0]).find('BLZ') + 5:str(data[0]).find('BLZ') + 13]
            for i in href:
                if i.get_text()[:4] == "http":
                    file2.write("{0};{1}\n".format(branch, i.get_text()))
                    continue
            # aprint(len(open("VrUrl.txt").readlines())-url,"left")
        except:
            print("Try this url:", url)
            continue
    file.close()
    file2.close()
    return 0


# Test
def get_data_sp_2(url, proxy=None):
    dataArraySp = np.array(["", "", 0, 0, 0], dtype="<U132")
    try:
        response = requests.get(url, proxy)
    except:
        print(url)
        return np.array([3, "wrong"])
    # In some cases multiple request attempts are necessary, thus failed urls are stored in an array for further usage
    if str(response) != "<Response [200]>":
        print(url)
        return np.array([2, "wrong"])

    soup = BeautifulSoup(response.text, "html.parser")
    data = soup.findAll('button')
    for i in data:
        try:
            return [i["data-blz"], i["value"]]
        except:
            1
    print(url)
    return [1, "wrong"]


list = ["st"]


def main_tag(proxy=None):
    file = open("UrlSpFiliale.txt", "r")
    CNT = 0
    cnt = 2
    wb = load_workbook("PostalcodeEXTRA.xlsx", read_only=False)  # , keep_vba=True)
    sheet = wb[wb.sheetnames[1]]
    lenUrl = len(open("UrlSpFiliale.txt").readlines())
    for iter in range(lenUrl):
        url = file.readline().replace('\n', '')
        # print(list[CNT]," - ",url)
        br = False
        for tag in list:
            if url[:50] in tag:
                br = True
        if br:
            CNT += 1
            continue

        data = get_data_sp2(url)
        data[1] = data[1].replace("https://www.", "").replace(".de/", "").replace(".de?n=true", "").replace(
            "http://www.", "").replace("/home.html", "").replace(".de", "")
        # print(data[1])
        list.append(url)
        # print(list)
        try:
            sheet.cell(row=cnt, column=1).value = float(data[0])
            sheet.cell(row=cnt, column=2).value = data[1]
            cnt += 1
            CNT += 1

        except:
            traceback.print_exc()
        # print(lenUrl-CNT)
        if "000" in str(CNT):
            print("save", lenUrl - CNT)
            wb.save("PostalcodeEXTRA.xlsx")

    wb.save("PostalcodeEXTRA.xlsx")
    file.close()
    return 0
